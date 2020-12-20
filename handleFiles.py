from openpyxl import Workbook, load_workbook

file1_write_instance = Workbook
file2_write_instance = Workbook
file3_write_instance = Workbook

file1_destination = "excel/new/appleStore_description_new.xlsx"
file2_destination = "excel/new/AppleStore_new.xlsx"
file3_destination = "excel/new/appleStore_additional_data.xlsx"

file_to_read = "excel/old/AppleStore.xlsx"

current_line_write_file = 3
count = 1

active_sheet_read_file = Workbook

active_sheet_write_file1 = Workbook
active_sheet_write_file2 = Workbook
active_sheet_write_file3 = Workbook


def get_active_sheet_read_file():
    return active_sheet_read_file


def write_data(data):
    global active_sheet_write_file1, active_sheet_write_file2, active_sheet_write_file3, current_line_write_file, count

    # file 1
    active_sheet_write_file1['A' + str(current_line_write_file)] = data["results"][0]["trackId"]
    active_sheet_write_file1['B' + str(current_line_write_file)] = data["results"][0]["trackName"]
    active_sheet_write_file1['C' + str(current_line_write_file)] = data["results"][0]["fileSizeBytes"]
    active_sheet_write_file1['D' + str(current_line_write_file)] = repr(data["results"][0]["description"])

    # file 2
    active_sheet_write_file2['A' + str(current_line_write_file)] = count
    active_sheet_write_file2['B' + str(current_line_write_file)] = data["results"][0]["trackId"]
    active_sheet_write_file2['C' + str(current_line_write_file)] = data["results"][0]["trackName"]
    active_sheet_write_file2['D' + str(current_line_write_file)] = data["results"][0]["fileSizeBytes"]
    active_sheet_write_file2['E' + str(current_line_write_file)] = data["results"][0]["currency"]
    active_sheet_write_file2['F' + str(current_line_write_file)] = data["results"][0]["price"]
    active_sheet_write_file2['G' + str(current_line_write_file)] = data["results"][0]["userRatingCount"]
    active_sheet_write_file2['H' + str(current_line_write_file)] = data["results"][0]["userRatingCountForCurrentVersion"]
    active_sheet_write_file2['I' + str(current_line_write_file)] = data["results"][0]["averageUserRating"]
    active_sheet_write_file2['J' + str(current_line_write_file)] = data["results"][0]["averageUserRatingForCurrentVersion"]
    active_sheet_write_file2['K' + str(current_line_write_file)] = data["results"][0]["version"]
    active_sheet_write_file2['L' + str(current_line_write_file)] = data["results"][0]["trackContentRating"]
    active_sheet_write_file2['M' + str(current_line_write_file)] = data["results"][0]["primaryGenreName"]
    active_sheet_write_file2['N' + str(current_line_write_file)] = len(data["results"][0]["supportedDevices"])
    active_sheet_write_file2['O' + str(current_line_write_file)] = len(data["results"][0]["ipadScreenshotUrls"])
    active_sheet_write_file2['P' + str(current_line_write_file)] = len(data["results"][0]["languageCodesISO2A"])
    active_sheet_write_file2['Q' + str(current_line_write_file)] = data["results"][0]["isVppDeviceBasedLicensingEnabled"]

    # file 3
    active_sheet_write_file3['A' + str(current_line_write_file)] = data["results"][0]["trackId"]
    active_sheet_write_file3['B' + str(current_line_write_file)] = data["results"][0]["minimumOsVersion"]
    if data["results"][0].get('releaseNotes'):
        active_sheet_write_file3['C' + str(current_line_write_file)] = repr(data["results"][0]["releaseNotes"])
    active_sheet_write_file3['D' + str(current_line_write_file)] = ', '.join(data["results"][0]["languageCodesISO2A"])
    active_sheet_write_file3['E' + str(current_line_write_file)] = ', '.join(data["results"][0]["supportedDevices"])
    active_sheet_write_file3['F' + str(current_line_write_file)] = len(data["results"][0]["screenshotUrls"])
    active_sheet_write_file3['G' + str(current_line_write_file)] = data["results"][0]["isGameCenterEnabled"]
    active_sheet_write_file3['H' + str(current_line_write_file)] = ', '.join(data["results"][0]["features"])
    active_sheet_write_file3['I' + str(current_line_write_file)] = data["results"][0]["releaseDate"]
    active_sheet_write_file3['J' + str(current_line_write_file)] = data["results"][0]["currentVersionReleaseDate"]
    active_sheet_write_file3['K' + str(current_line_write_file)] = ', '.join(data["results"][0]["genres"])
    active_sheet_write_file3['L' + str(current_line_write_file)] = data["results"][0]["formattedPrice"]



    count += 1
    current_line_write_file += 1


def get_active_sheet_write_file():
    return active_sheet_write_file1


def initialise():
    global active_sheet_read_file
    wb1 = load_workbook(filename=file_to_read)
    active_sheet_read_file = wb1.active

    global active_sheet_write_file1, file1_write_instance
    file1_write_instance = Workbook()
    active_sheet_write_file1 = file1_write_instance.active

    global active_sheet_write_file2, file2_write_instance
    file2_write_instance = Workbook()
    active_sheet_write_file2 = file2_write_instance.active

    global active_sheet_write_file3, file3_write_instance
    file3_write_instance = Workbook()
    active_sheet_write_file3 = file3_write_instance.active


def save_files():
    global file3_write_instance, active_sheet_write_file3
    file1_write_instance.save(filename=file1_destination)
    file2_write_instance.save(filename=file2_destination)
    file3_write_instance.save(filename=file3_destination)
